from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Count, Q, Avg, Sum, Max, Min
from django.contrib import messages
from django.urls import reverse
from datetime import date, timedelta
import json
import uuid
from .models import Reporte, TipoReporte, MetricasRendimiento

@login_required
def dashboard_reportes_view(request):
    """Dashboard principal de reportes con estadísticas y lista completa"""
    try:
        # Filtros
        estado_filtro = request.GET.get('estado', '')
        tipo_filtro = request.GET.get('tipo', '')
        busqueda = request.GET.get('busqueda', '')

        # Estadísticas de reportes
        total_reportes = Reporte.objects.count()
        reportes_mes = Reporte.objects.filter(
            fecha_solicitud__month=timezone.now().month,
            fecha_solicitud__year=timezone.now().year
        ).count()
        tipos_disponibles = TipoReporte.objects.filter(activo=True).count()
        reportes_pendientes = Reporte.objects.filter(estado='pendiente').count()

        # Query base para lista de reportes
        reportes = Reporte.objects.select_related(
            'tipo_reporte', 'usuario_solicitante'
        ).order_by('-fecha_solicitud')

        # Aplicar filtros
        if estado_filtro:
            reportes = reportes.filter(estado=estado_filtro)

        if tipo_filtro:
            reportes = reportes.filter(tipo_reporte_id=tipo_filtro)

        if busqueda:
            reportes = reportes.filter(
                Q(titulo__icontains=busqueda) |
                Q(descripcion__icontains=busqueda) |
                Q(tipo_reporte__nombre__icontains=busqueda)
            )

        # Paginación
        paginator = Paginator(reportes, 15)  # 15 reportes por página
        page_number = request.GET.get('page')
        reportes_page = paginator.get_page(page_number)

        # Datos para filtros
        tipos_reporte = TipoReporte.objects.filter(activo=True)
        estados = Reporte.ESTADO_CHOICES

        context = {
            'title': 'Dashboard Reportes',
            'total_reportes': total_reportes,
            'reportes_mes': reportes_mes,
            'tipos_disponibles': tipos_disponibles,
            'reportes_pendientes': reportes_pendientes,
            'reportes': reportes_page,
            'tipos_reporte': tipos_reporte,
            'estados': estados,
            'estado_filtro': estado_filtro,
            'tipo_filtro': tipo_filtro,
            'busqueda': busqueda,
        }

    except Exception as e:
        messages.error(request, f'Error al cargar dashboard: {str(e)}')
        context = {'title': 'Dashboard Reportes', 'reportes': []}

    return render(request, 'reportes/dashboard_reportes.html', context)

@login_required
def visual_dashboard_view(request):
    return render(request, 'reportes/visual_dashboard.html', {'title': 'Dashboard Visual'})

@login_required
def generar_reporte_view(request):
    """Vista para mostrar formulario de generar reporte (solo GET)"""
    try:
        from maquinaria.models import CategoriaMaquina, Maquina
        from documentos.models import Documento

        tipos_reporte = TipoReporte.objects.filter(activo=True)
        categorias = CategoriaMaquina.objects.filter(activa=True)
        from usuarios.models import Usuario as UsuarioModel
        centros_raw = set(
            UsuarioModel.objects.exclude(centro_formacion__isnull=True)
            .values_list('centro_formacion__nombre', flat=True)
        ) | set(
            Maquina.objects.values_list('centro_formacion', flat=True)
            .exclude(centro_formacion='').exclude(centro_formacion__isnull=True)
        )
        # Deduplicar ignorando mayúsculas/minúsculas: mantener la versión .title()
        seen = {}
        for c in centros_raw:
            key = c.lower()
            if key not in seen:
                seen[key] = c.title()
        centros_formacion = sorted(seen.values())

        # Rango real de datos (maquinaria + documentos, sin backups)
        maq_dates = Maquina.objects.aggregate(
            min_d=Min('fecha_adquisicion'),
            max_d=Max('fecha_adquisicion')
        )
        doc_dates = Documento.objects.aggregate(
            min_d=Min('fecha_creacion'),
            max_d=Max('fecha_creacion')
        )

        fechas_min = [
            d.date() if hasattr(d, 'date') else d
            for d in [maq_dates['min_d'], doc_dates['min_d']] if d
        ]
        fechas_max = [
            d.date() if hasattr(d, 'date') else d
            for d in [maq_dates['max_d'], doc_dates['max_d']] if d
        ]

        fecha_datos_inicio = min(fechas_min) if fechas_min else None
        fecha_datos_fin    = max(fechas_max) if fechas_max else date.today()

        context = {
            'title': 'Generar Reporte',
            'tipos_reporte': tipos_reporte,
            'categorias': categorias,
            'centros_formacion': centros_formacion,
            'fecha_datos_inicio': fecha_datos_inicio,
            'fecha_datos_fin': fecha_datos_fin,
        }

    except ImportError:
        context = {'title': 'Generar Reporte'}

    return render(request, 'reportes/generar_reporte.html', context)

# ---------------------------------------------------------------------------
# Helpers para recolectar datos por tipo de reporte
# Cada función devuelve una lista de secciones:
#   [{'titulo': str, 'headers': [...], 'rows': [[...], ...]}, ...]
# ---------------------------------------------------------------------------

def _filtrar_maquinas(reporte):
    from maquinaria.models import Maquina
    qs = Maquina.objects.select_related('categoria', 'proveedor').all()
    if reporte.categorias_maquina:
        qs = qs.filter(categoria_id__in=reporte.categorias_maquina)
    if reporte.centros_formacion:
        qs = qs.filter(centro_formacion__in=reporte.centros_formacion)
    if reporte.estados_maquina:
        qs = qs.filter(estado__in=reporte.estados_maquina)
    if reporte.fecha_inicio:
        qs = qs.filter(fecha_adquisicion__gte=reporte.fecha_inicio)
    if reporte.fecha_fin:
        qs = qs.filter(fecha_adquisicion__lte=reporte.fecha_fin)
    return qs


def _datos_reporte_maquinaria(reporte, with_details=False):
    from maquinaria.models import MantenimientoProgramado, UsoMaquinaria
    maquinas = _filtrar_maquinas(reporte)

    # Sección 1: Maquinaria general
    h_maq = ['Código', 'Nombre', 'Categoría', 'Marca', 'Modelo', 'Estado',
              'Condición', 'Eficiencia (%)', 'Horas Uso Total', 'Centro',
              'Ubicación', 'Fecha Adquisición', 'Valor Adquisición']
    if with_details:
        h_maq += ['N° Serie', 'Potencia', 'Voltaje', 'Dimensiones', 'Peso',
                  'Garantía (meses)', 'Horas Uso Mes', 'Ambiente Formación']
    r_maq = []
    ids_maquinas = []
    for m in maquinas:
        ids_maquinas.append(m.id)
        row = [
            m.codigo_inventario, m.nombre,
            m.categoria.nombre if m.categoria else 'N/A',
            m.marca, m.modelo,
            m.get_estado_display(), m.get_condicion_display(),
            str(m.eficiencia), str(m.horas_uso_total),
            m.centro_formacion, m.ubicacion,
            str(m.fecha_adquisicion), str(m.valor_adquisicion),
        ]
        if with_details:
            row += [
                m.numero_serie, m.potencia or 'N/A', m.voltaje or 'N/A',
                m.dimensiones or 'N/A', m.peso or 'N/A',
                str(m.garantia_meses), str(m.horas_uso_mes),
                m.ambiente_formacion or 'N/A',
            ]
        r_maq.append(row)

    # Sección 2: Mantenimientos
    h_mant = ['Máquina', 'Tipo', 'Título', 'Estado', 'Prioridad',
              'Fecha Programada', 'Fecha Completado', 'Costo Real', 'Técnico']
    r_mant = []
    mants = MantenimientoProgramado.objects.filter(
        maquina_id__in=ids_maquinas
    ).select_related('maquina', 'tecnico_realizado').order_by('-fecha_programada')
    if reporte.fecha_inicio:
        mants = mants.filter(fecha_programada__date__gte=reporte.fecha_inicio)
    if reporte.fecha_fin:
        mants = mants.filter(fecha_programada__date__lte=reporte.fecha_fin)
    for mt in mants:
        r_mant.append([
            mt.maquina.codigo_inventario,
            mt.get_tipo_display(), mt.titulo,
            mt.get_estado_display(), mt.get_prioridad_display(),
            mt.fecha_programada.strftime('%Y-%m-%d %H:%M') if mt.fecha_programada else 'N/A',
            mt.fecha_fin_real.strftime('%Y-%m-%d %H:%M') if mt.fecha_fin_real else 'N/A',
            str(mt.costo_real) if mt.costo_real else 'N/A',
            str(mt.tecnico_realizado) if mt.tecnico_realizado else 'N/A',
        ])

    # Sección 3: Usos registrados
    h_uso = ['Máquina', 'Fecha', 'Ficha', 'Hora Inicio', 'Hora Fin',
             'Horas Sesión', 'Horas Totales Máquina', 'Actividad', 'Instructor']
    r_uso = []
    usos = UsoMaquinaria.objects.filter(
        maquina_id__in=ids_maquinas
    ).select_related('maquina', 'instructor_encargado').order_by('-fecha')
    if reporte.fecha_inicio:
        usos = usos.filter(fecha__gte=reporte.fecha_inicio)
    if reporte.fecha_fin:
        usos = usos.filter(fecha__lte=reporte.fecha_fin)
    for u in usos:
        r_uso.append([
            u.maquina.codigo_inventario,
            str(u.fecha), u.ficha,
            str(u.hora_inicio), str(u.hora_fin),
            str(u.horas_uso), str(u.horas_totales_maquina),
            u.descripcion_actividad[:80],
            str(u.instructor_encargado) if u.instructor_encargado else 'N/A',
        ])

    return [
        {'titulo': 'Maquinaria', 'headers': h_maq, 'rows': r_maq},
        {'titulo': 'Mantenimientos Realizados', 'headers': h_mant, 'rows': r_mant},
        {'titulo': 'Usos Registrados', 'headers': h_uso, 'rows': r_uso},
    ]


def _datos_reporte_inventario(reporte, with_details=False):
    from inventario.models import PiezaInventario
    qs = PiezaInventario.objects.select_related('responsable').order_by('-fecha_registro')
    if reporte.centros_formacion:
        qs = qs.filter(centro_formacion__in=reporte.centros_formacion)
    if reporte.fecha_inicio:
        qs = qs.filter(fecha_registro__date__gte=reporte.fecha_inicio)
    if reporte.fecha_fin:
        qs = qs.filter(fecha_registro__date__lte=reporte.fecha_fin)

    h = ['Código', 'Nombre', 'Categoría', 'Marca', 'Modelo', 'Condición',
         'Horas Uso', 'Horas Máx.', 'Fecha Compra', 'Fecha Registro',
         'Centro', 'Ubicación', 'Valor Adquisición', 'Responsable']
    if with_details:
        h += ['N° Serie', 'N° Factura', 'Fabricante', 'Color',
              'Dimensiones', 'Peso', 'Garantía (meses)', 'Observaciones']
    rows = []
    for p in qs:
        row = [
            p.codigo_inventario, p.nombre,
            p.categoria or 'N/A', p.marca or 'N/A', p.modelo or 'N/A',
            p.get_condicion_display(),
            str(p.horas_uso),
            str(p.horas_uso_maximas) if p.horas_uso_maximas else 'N/A',
            str(p.fecha_compra) if p.fecha_compra else 'N/A',
            p.fecha_registro.strftime('%Y-%m-%d') if p.fecha_registro else 'N/A',
            p.centro_formacion or 'N/A', p.ubicacion or 'N/A',
            str(p.valor_adquisicion) if p.valor_adquisicion else 'N/A',
            str(p.responsable) if p.responsable else 'N/A',
        ]
        if with_details:
            row += [
                p.numero_serie or 'N/A', p.numero_factura or 'N/A',
                p.fabricante or 'N/A', p.color or 'N/A',
                p.dimensiones or 'N/A', p.peso or 'N/A',
                str(p.garantia_meses), p.observaciones[:80] if p.observaciones else 'N/A',
            ]
        rows.append(row)
    return [{'titulo': 'Inventario de Piezas', 'headers': h, 'rows': rows}]


def _datos_reporte_documentos(reporte, with_details=False):
    from documentos.models import Documento, TipoDocumento, CategoriaDocumento
    qs = Documento.objects.select_related('tipo_documento', 'categoria', 'creado_por').all()
    if reporte.fecha_inicio:
        qs = qs.filter(fecha_creacion__date__gte=reporte.fecha_inicio)
    if reporte.fecha_fin:
        qs = qs.filter(fecha_creacion__date__lte=reporte.fecha_fin)

    # Resumen por Tipo
    por_tipo = qs.values('tipo_documento__nombre').annotate(total=Count('id')).order_by('-total')
    h_tipo = ['Tipo de Documento', 'Cantidad']
    r_tipo = [[t['tipo_documento__nombre'] or 'N/A', str(t['total'])] for t in por_tipo]

    # Resumen por Categoría
    por_cat = qs.values('categoria__nombre').annotate(total=Count('id')).order_by('-total')
    h_cat = ['Categoría', 'Cantidad']
    r_cat = [[c['categoria__nombre'] or 'N/A', str(c['total'])] for c in por_cat]

    # Resumen por Estado
    por_estado = qs.values('estado').annotate(total=Count('id')).order_by('-total')
    estado_map = dict(Documento.ESTADO_CHOICES)
    h_est = ['Estado', 'Cantidad']
    r_est = [[estado_map.get(e['estado'], e['estado']), str(e['total'])] for e in por_estado]

    # Resumen por Nivel de Acceso
    por_acceso = qs.values('nivel_acceso').annotate(total=Count('id')).order_by('-total')
    acceso_map = dict(Documento.NIVEL_ACCESO_CHOICES)
    h_acc = ['Nivel de Acceso', 'Cantidad']
    r_acc = [[acceso_map.get(a['nivel_acceso'], a['nivel_acceso']), str(a['total'])] for a in por_acceso]

    # Listado detallado
    h_det = ['Título', 'Tipo', 'Categoría', 'Estado', 'Nivel Acceso',
             'Versión', 'Fecha Creación', 'Descargas', 'Creado Por']
    if with_details:
        h_det += ['Visualizaciones', 'Autor Original', 'Palabras Clave',
                  'Fecha Publicación', 'Fecha Vencimiento']
    r_det = []
    for d in qs.order_by('-fecha_creacion'):
        row = [
            d.titulo,
            d.tipo_documento.nombre if d.tipo_documento else 'N/A',
            d.categoria.nombre if d.categoria else 'N/A',
            dict(Documento.ESTADO_CHOICES).get(d.estado, d.estado),
            dict(Documento.NIVEL_ACCESO_CHOICES).get(d.nivel_acceso, d.nivel_acceso),
            d.version,
            d.fecha_creacion.strftime('%Y-%m-%d') if d.fecha_creacion else 'N/A',
            str(d.total_descargas),
            str(d.creado_por) if d.creado_por else 'N/A',
        ]
        if with_details:
            row += [
                str(d.total_visualizaciones),
                d.autor_original or 'N/A',
                ', '.join(d.palabras_clave) if d.palabras_clave else 'N/A',
                d.fecha_publicacion.strftime('%Y-%m-%d') if d.fecha_publicacion else 'N/A',
                str(d.fecha_vencimiento) if d.fecha_vencimiento else 'N/A',
            ]
        r_det.append(row)

    return [
        {'titulo': 'Por Tipo de Documento', 'headers': h_tipo, 'rows': r_tipo},
        {'titulo': 'Por Categoría', 'headers': h_cat, 'rows': r_cat},
        {'titulo': 'Por Estado', 'headers': h_est, 'rows': r_est},
        {'titulo': 'Por Nivel de Acceso', 'headers': h_acc, 'rows': r_acc},
        {'titulo': 'Listado Detallado', 'headers': h_det, 'rows': r_det},
    ]


def _generar_archivo_reporte(reporte):
    """Genera el archivo del reporte en el formato solicitado y lo guarda en reporte.archivo_resultado"""
    import io
    import csv as csv_mod
    from django.core.files.base import ContentFile

    try:
        params = reporte.parametros or {}
        with_graficos = params.get('incluir_graficos', False)
        with_detalles = params.get('incluir_detalles', False)
        send_email    = params.get('notificar_email', False)

        tipo_nombre = reporte.tipo_reporte.nombre.lower()
        fmt = reporte.formato.lower()

        if 'maquinaria' in tipo_nombre:
            secciones = _datos_reporte_maquinaria(reporte, with_details=with_detalles)
        elif 'inventario' in tipo_nombre:
            secciones = _datos_reporte_inventario(reporte, with_details=with_detalles)
        elif 'documento' in tipo_nombre:
            secciones = _datos_reporte_documentos(reporte, with_details=with_detalles)
        else:
            secciones = _datos_reporte_maquinaria(reporte, with_details=with_detalles)

        total_registros = sum(len(s['rows']) for s in secciones)

        if fmt == 'csv':
            buf = io.StringIO()
            writer = csv_mod.writer(buf)
            for seccion in secciones:
                writer.writerow([f'=== {seccion["titulo"]} ==='])
                writer.writerow(seccion['headers'])
                writer.writerows(seccion['rows'])
                writer.writerow([])
            content = buf.getvalue().encode('utf-8-sig')
            filename = f'reporte_{reporte.id}.csv'
            file_obj = ContentFile(content, name=filename)

        elif fmt == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # eliminar hoja vacía por defecto
            sena_green = 'FF39a900'
            sena_blue = 'FF00324d'
            alt_row = 'FFf0f8e8'

            for seccion in secciones:
                ws = wb.create_sheet(title=seccion['titulo'][:31])
                headers = seccion['headers']
                rows = seccion['rows']
                last_col_letter = chr(64 + len(headers)) if len(headers) <= 26 else 'Z'

                # Título de la sección
                ws.merge_cells(f'A1:{last_col_letter}1')
                ws['A1'] = f"{reporte.titulo} — {seccion['titulo']}"
                ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
                ws['A1'].fill = PatternFill('solid', fgColor=sena_blue)
                ws['A1'].alignment = Alignment(horizontal='center')

                # Cabeceras
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color='FFFFFF', size=9)
                    cell.fill = PatternFill('solid', fgColor=sena_green)
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)

                # Datos
                for row_idx, row in enumerate(rows, 3):
                    fill = PatternFill('solid', fgColor=alt_row) if row_idx % 2 == 0 else None
                    for col_idx, val in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        if fill:
                            cell.fill = fill

                # Ajustar anchos
                for col in ws.columns:
                    max_len = max((len(str(c.value or '')) for c in col), default=8)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

            # Gráficos en Excel
            if with_graficos:
                from openpyxl.chart import BarChart, PieChart, Reference
                ws_graf = wb.create_sheet(title='Gráficos', index=0)
                ws_graf['A1'] = 'Gráficos del Reporte'
                ws_graf['A1'].font = Font(bold=True, size=14, color='FFFFFF')
                ws_graf['A1'].fill = PatternFill('solid', fgColor=sena_blue)

                # Tomar primera sección con datos para el gráfico de barras
                for seccion in secciones:
                    if len(seccion['rows']) > 0:
                        # Datos auxiliares en ws_graf para el gráfico
                        ws_graf['A3'] = seccion['titulo']
                        ws_graf['A3'].font = Font(bold=True)
                        ws_graf['B3'] = 'Cantidad'
                        # Contar por primera columna (agrupado)
                        conteo = {}
                        for row in seccion['rows']:
                            k = str(row[0])[:20]
                            conteo[k] = conteo.get(k, 0) + 1
                        for i, (k, v) in enumerate(list(conteo.items())[:15], 4):
                            ws_graf.cell(row=i, column=1, value=k)
                            ws_graf.cell(row=i, column=2, value=v)

                        n = min(len(conteo), 15)
                        if n > 0:
                            bar = BarChart()
                            bar.type = 'col'
                            bar.title = seccion['titulo']
                            bar.y_axis.title = 'Cantidad'
                            bar.style = 10
                            bar.width = 20
                            bar.height = 12
                            data_ref = Reference(ws_graf, min_col=2, min_row=3, max_row=3 + n)
                            cats_ref = Reference(ws_graf, min_col=1, min_row=4, max_row=3 + n)
                            bar.add_data(data_ref, titles_from_data=True)
                            bar.set_categories(cats_ref)
                            ws_graf.add_chart(bar, 'D3')
                        break

            buf = io.BytesIO()
            wb.save(buf)
            content = buf.getvalue()
            filename = f'reporte_{reporte.id}.xlsx'
            file_obj = ContentFile(content, name=filename)

        elif fmt == 'pdf':
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Paragraph, Spacer, PageBreak)
            from reportlab.lib.units import cm

            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                    topMargin=1.5*cm, bottomMargin=1.5*cm,
                                    leftMargin=1.5*cm, rightMargin=1.5*cm)
            styles = getSampleStyleSheet()
            sena_green = colors.HexColor('#39a900')
            sena_blue = colors.HexColor('#00324d')
            alt_color = colors.HexColor('#f0f8e8')

            title_style = ParagraphStyle('rtitle', parent=styles['Title'],
                                         textColor=sena_blue, fontSize=15, spaceAfter=6)
            section_style = ParagraphStyle('rsec', parent=styles['Heading2'],
                                           textColor=colors.white, fontSize=11,
                                           backColor=sena_blue, spaceAfter=4,
                                           leftIndent=6, rightIndent=6)
            meta_style = ParagraphStyle('rmeta', parent=styles['Normal'],
                                        fontSize=8, textColor=colors.grey, spaceAfter=10)

            story = []
            story.append(Paragraph(reporte.titulo, title_style))
            story.append(Paragraph(
                f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')} | "
                f"Tipo: {reporte.tipo_reporte.nombre} | "
                f"Total registros: {total_registros}",
                meta_style
            ))

            # Gráficos en PDF
            if with_graficos:
                from reportlab.graphics.shapes import Drawing
                from reportlab.graphics.charts.barcharts import VerticalBarChart
                from reportlab.graphics.charts.piecharts import Pie
                from reportlab.graphics import renderPDF

                grafico_style = ParagraphStyle('rgraf', parent=styles['Heading3'],
                                               textColor=sena_blue, spaceAfter=4, spaceBefore=12)
                CHART_COLORS = [
                    colors.HexColor('#39a900'), colors.HexColor('#00324d'),
                    colors.HexColor('#ff6b35'), colors.HexColor('#ffc107'),
                    colors.HexColor('#17a2b8'), colors.HexColor('#6f42c1'),
                    colors.HexColor('#dc3545'), colors.HexColor('#fd7e14'),
                ]

                for seccion in secciones:
                    rows = seccion['rows']
                    if not rows:
                        continue
                    # Conteo por primera columna para barras
                    conteo = {}
                    for row in rows:
                        k = str(row[0])[:25]
                        conteo[k] = conteo.get(k, 0) + 1
                    keys = list(conteo.keys())[:8]
                    vals = [conteo[k] for k in keys]
                    if not vals:
                        continue

                    story.append(Paragraph(f'Distribución — {seccion["titulo"]}', grafico_style))

                    # Barra vertical
                    d = Drawing(500, 160)
                    bc = VerticalBarChart()
                    bc.x, bc.y, bc.width, bc.height = 50, 10, 430, 130
                    bc.data = [vals]
                    bc.categoryAxis.categoryNames = keys
                    bc.categoryAxis.labels.angle = 20
                    bc.categoryAxis.labels.fontSize = 6
                    bc.valueAxis.valueMin = 0
                    bc.valueAxis.valueMax = max(vals) + 1
                    bc.bars[0].fillColor = sena_green
                    bc.bars[0].strokeColor = colors.white
                    d.add(bc)
                    story.append(d)

                    # Pie chart
                    d2 = Drawing(300, 160)
                    pie = Pie()
                    pie.x, pie.y, pie.width, pie.height = 50, 10, 140, 140
                    pie.data = vals
                    pie.labels = [f'{k[:15]}({v})' for k, v in zip(keys, vals)]
                    pie.sideLabels = True
                    for i, color in enumerate(CHART_COLORS[:len(vals)]):
                        pie.slices[i].fillColor = color
                        pie.slices[i].strokeColor = colors.white
                        pie.slices[i].strokeWidth = 0.5
                    pie.slices.labelRadius = 1.3
                    pie.slices.fontSize = 6
                    d2.add(pie)
                    story.append(d2)
                    story.append(Spacer(1, 0.4*cm))

            page_w = landscape(A4)[0] - 3*cm

            for i, seccion in enumerate(secciones):
                if i > 0:
                    story.append(Spacer(1, 0.6*cm))
                story.append(Paragraph(seccion['titulo'], section_style))
                story.append(Spacer(1, 0.2*cm))

                headers = seccion['headers']
                rows = seccion['rows']
                if not rows:
                    story.append(Paragraph('Sin registros en este período.', styles['Normal']))
                    continue

                col_w = page_w / len(headers)
                table_data = [headers] + rows
                t = Table(table_data, colWidths=[col_w] * len(headers), repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), sena_green),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 7),
                    ('FONTSIZE', (0, 1), (-1, -1), 6.5),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_color]),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
                story.append(t)

            doc.build(story)
            content = buf.getvalue()
            filename = f'reporte_{reporte.id}.pdf'
            file_obj = ContentFile(content, name=filename)

        else:  # json
            payload = {
                'titulo': reporte.titulo,
                'tipo': reporte.tipo_reporte.nombre,
                'generado': timezone.now().isoformat(),
                'total_registros': total_registros,
                'secciones': [
                    {
                        'titulo': s['titulo'],
                        'columnas': s['headers'],
                        'datos': [dict(zip(s['headers'], row)) for row in s['rows']],
                    }
                    for s in secciones
                ],
            }
            content = json.dumps(payload, ensure_ascii=False, indent=2).encode('utf-8')
            filename = f'reporte_{reporte.id}.json'
            file_obj = ContentFile(content, name=filename)

        reporte.archivo_resultado.save(filename, file_obj, save=False)
        reporte.total_registros = total_registros
        reporte.tamaño_archivo = len(content)
        reporte.estado = 'completado'
        reporte.fecha_completado = timezone.now()
        reporte.save()

        # Notificación por email
        if send_email:
            try:
                from django.core.mail import EmailMessage as DjangoEmail
                from django.conf import settings

                destinatario = reporte.usuario_solicitante.email
                if destinatario:
                    ext_map = {'pdf': 'application/pdf',
                               'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               'csv': 'text/csv',
                               'json': 'application/json'}
                    mime = ext_map.get(fmt, 'application/octet-stream')

                    cuerpo = (
                        f'Hola {reporte.usuario_solicitante.nombres},\n\n'
                        f'Tu reporte "{reporte.titulo}" ha sido generado exitosamente.\n\n'
                        f'• Tipo: {reporte.tipo_reporte.nombre}\n'
                        f'• Formato: {fmt.upper()}\n'
                        f'• Total de registros: {total_registros}\n'
                        f'• Generado el: {reporte.fecha_completado.strftime("%d/%m/%Y %H:%M")}\n\n'
                        f'El archivo está adjunto a este correo.\n\n'
                        f'SENA – Sistema de Gestión de Maquinaria'
                    )
                    email = DjangoEmail(
                        subject=f'Reporte listo: {reporte.titulo}',
                        body=cuerpo,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[destinatario],
                    )
                    email.attach(filename, content, mime)
                    email.send(fail_silently=True)
            except Exception:
                pass  # El email falla silenciosamente, el reporte ya fue generado

    except Exception as e:
        reporte.estado = 'error'
        reporte.error_mensaje = str(e)
        reporte.save()
        raise


@login_required
@require_http_methods(["POST"])
def crear_reporte_web(request):
    """Endpoint POST para crear reporte desde formulario web"""
    try:
        from usuarios.models import Usuario

        tipo_reporte_id = request.POST.get('tipo_reporte')
        nombre_reporte = request.POST.get('nombre_reporte')
        descripcion = request.POST.get('descripcion', '')
        formato = request.POST.get('formato', 'pdf')

        if not tipo_reporte_id:
            messages.error(request, 'El tipo de reporte es requerido')
            return redirect('reportes:generar_reporte')

        if not nombre_reporte:
            messages.error(request, 'El nombre del reporte es requerido')
            return redirect('reportes:generar_reporte')

        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        categoria = request.POST.getlist('categoria')
        centro_formacion = request.POST.get('centro_formacion')
        estado_maquina = request.POST.get('estado_maquina')

        # Opciones del reporte
        incluir_graficos    = 'incluir_graficos' in request.POST
        incluir_detalles    = 'incluir_detalles' in request.POST
        generar_automatico  = 'generar_automatico' in request.POST
        notificar_email     = 'notificar_email' in request.POST

        try:
            tipo_reporte = TipoReporte.objects.get(id=tipo_reporte_id)

            if not request.user.is_authenticated:
                messages.error(request, 'Debes estar autenticado para crear reportes')
                return redirect('usuarios:login')

            try:
                usuario_solicitante = Usuario.objects.get(numero_documento=request.user.username)
            except Usuario.DoesNotExist:
                messages.error(request, 'Tu perfil de usuario no está completo. Contacta al administrador.')
                return redirect('reportes:generar_reporte')

            reporte = Reporte.objects.create(
                tipo_reporte=tipo_reporte,
                usuario_solicitante=usuario_solicitante,
                titulo=nombre_reporte,
                descripcion=descripcion,
                formato=formato,
                fecha_inicio=fecha_inicio if fecha_inicio else None,
                fecha_fin=fecha_fin if fecha_fin else None,
                categorias_maquina=categoria,
                centros_formacion=[centro_formacion] if centro_formacion else [],
                estados_maquina=[estado_maquina] if estado_maquina else [],
                parametros={
                    'incluir_graficos': incluir_graficos,
                    'incluir_detalles': incluir_detalles,
                    'generar_automatico': generar_automatico,
                    'notificar_email': notificar_email,
                },
                estado='pendiente' if generar_automatico else 'generando'
            )

            if generar_automatico:
                messages.info(request, f'Reporte "{nombre_reporte}" programado. Se generará automáticamente en el siguiente ciclo de procesamiento.')
                return redirect('reportes:detalle_reporte', pk=reporte.id)

            try:
                _generar_archivo_reporte(reporte)
                messages.success(request, f'Reporte "{nombre_reporte}" generado exitosamente.')
            except Exception as e:
                messages.warning(request, f'Reporte creado pero con error al generar: {str(e)}')

            return redirect('reportes:detalle_reporte', pk=reporte.id)

        except TipoReporte.DoesNotExist:
            messages.error(request, 'Tipo de reporte no válido')
            return redirect('reportes:generar_reporte')

    except Exception as e:
        messages.error(request, f'Error al crear reporte: {str(e)}')
        return redirect('reportes:generar_reporte')

@csrf_exempt
@require_http_methods(["POST"])
def crear_reporte_api(request):
    """Endpoint POST API para crear un nuevo reporte"""
    try:
        from usuarios.models import Usuario

        # Verificar Content-Type
        if request.content_type == 'application/json':
            import json
            data = json.loads(request.body)
        else:
            data = request.POST

        # Obtener datos del reporte
        tipo_reporte_id = data.get('tipo_reporte')
        nombre_reporte = data.get('nombre_reporte')
        descripcion = data.get('descripcion', '')
        formato = data.get('formato', 'pdf')
        usuario_id = data.get('usuario_id')  # Para permitir especificar usuario

        # Validaciones básicas
        if not tipo_reporte_id:
            return JsonResponse({
                'success': False,
                'error': 'El tipo de reporte es requerido'
            }, status=400)

        if not nombre_reporte:
            return JsonResponse({
                'success': False,
                'error': 'El nombre del reporte es requerido'
            }, status=400)

        # Obtener filtros
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        categoria = data.get('categoria')
        centro_formacion = data.get('centro_formacion')
        estado_maquina = data.get('estado_maquina')

        try:
            tipo_reporte = TipoReporte.objects.get(id=tipo_reporte_id)

            # Determinar usuario solicitante
            if usuario_id:
                try:
                    usuario_solicitante = Usuario.objects.get(id=usuario_id)
                except Usuario.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'Usuario especificado no encontrado'
                    }, status=400)
            elif request.user.is_authenticated:
                # Si el usuario está autenticado, usar su instancia de Usuario
                try:
                    usuario_solicitante = Usuario.objects.get(user=request.user)
                except Usuario.DoesNotExist:
                    # Si no existe Usuario para este User, crear uno básico o usar el primer usuario
                    usuario_solicitante = Usuario.objects.first()
                    if not usuario_solicitante:
                        return JsonResponse({
                            'success': False,
                            'error': 'No hay usuarios en el sistema'
                        }, status=400)
            else:
                # Para testing sin autenticación, usar el primer usuario disponible
                usuario_solicitante = Usuario.objects.first()
                if not usuario_solicitante:
                    return JsonResponse({
                        'success': False,
                        'error': 'No hay usuarios en el sistema para asignar el reporte'
                    }, status=400)

            # Crear el reporte
            reporte = Reporte.objects.create(
                tipo_reporte=tipo_reporte,
                usuario_solicitante=usuario_solicitante,
                titulo=nombre_reporte,
                descripcion=descripcion,
                formato=formato,
                fecha_inicio=fecha_inicio if fecha_inicio else None,
                fecha_fin=fecha_fin if fecha_fin else None,
                categorias_maquina=[categoria] if categoria else [],
                centros_formacion=[centro_formacion] if centro_formacion else [],
                estados_maquina=[estado_maquina] if estado_maquina else [],
                estado='pendiente'
            )

            # Devolver JSON (es API)
            return JsonResponse({
                'success': True,
                'message': f'Reporte "{nombre_reporte}" creado exitosamente',
                'reporte_id': str(reporte.id),
                'reporte': {
                    'id': str(reporte.id),
                    'titulo': reporte.titulo,
                    'descripcion': reporte.descripcion,
                    'estado': reporte.estado,
                    'formato': reporte.formato,
                    'fecha_solicitud': reporte.fecha_solicitud.isoformat(),
                    'tipo_reporte': reporte.tipo_reporte.nombre,
                    'usuario': reporte.usuario_solicitante.nombre_completo if hasattr(reporte.usuario_solicitante, 'nombre_completo') else str(reporte.usuario_solicitante)
                }
            }, status=201)

        except TipoReporte.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Tipo de reporte no válido'
            }, status=400)

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'JSON inválido'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }, status=500)

@login_required
def lista_reportes_view(request):
    """Vista que redirige al dashboard (la lista ahora está en el dashboard)"""
    # Preservar los parámetros de búsqueda si existen
    query_string = request.GET.urlencode()
    if query_string:
        return redirect(f"{reverse('reportes:dashboard')}?{query_string}")
    return redirect('reportes:dashboard')

@login_required
def detalle_reporte_view(request, pk):
    """Vista para ver detalle de un reporte"""
    try:
        from usuarios.models import Usuario

        reporte = get_object_or_404(Reporte, id=pk)

        # Verificar que el usuario puede ver este reporte
        # Obtener el Usuario basado en el request.user.username (numero_documento)
        try:
            usuario_actual = Usuario.objects.get(numero_documento=request.user.username)
            if reporte.usuario_solicitante != usuario_actual and not request.user.is_staff:
                messages.error(request, 'No tienes permiso para ver este reporte')
                return redirect('reportes:lista_reportes')
        except Usuario.DoesNotExist:
            # Si no existe el usuario, solo permitir a staff
            if not request.user.is_staff:
                messages.error(request, 'No tienes permiso para ver este reporte')
                return redirect('reportes:lista_reportes')

        context = {
            'title': f'Reporte: {reporte.titulo}',
            'reporte': reporte,
        }

    except Exception as e:
        messages.error(request, f'Error al cargar reporte: {str(e)}')
        return redirect('reportes:lista_reportes')

    return render(request, 'reportes/detalle_reporte.html', context)

@login_required
def descargar_reporte(request, pk):
    """Sirve el archivo del reporte para descarga"""
    try:
        from usuarios.models import Usuario
        reporte = get_object_or_404(Reporte, id=pk)

        try:
            usuario_actual = Usuario.objects.get(numero_documento=request.user.username)
            if reporte.usuario_solicitante != usuario_actual and not request.user.is_staff:
                raise Http404
        except Usuario.DoesNotExist:
            if not request.user.is_staff:
                raise Http404

        if not reporte.archivo_resultado:
            messages.error(request, 'El archivo del reporte no está disponible.')
            return redirect('reportes:detalle_reporte', pk=pk)

        import mimetypes
        file_path = reporte.archivo_resultado.path
        mime_type, _ = mimetypes.guess_type(file_path)
        mime_type = mime_type or 'application/octet-stream'

        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=mime_type)

        response['Content-Disposition'] = (
            f'attachment; filename="{reporte.titulo.replace(" ", "_")}'
            f'.{reporte.formato}"'
        )
        reporte.veces_descargado += 1
        reporte.fecha_ultima_descarga = timezone.now()
        reporte.save(update_fields=['veces_descargado', 'fecha_ultima_descarga'])
        return response

    except Http404:
        raise
    except Exception as e:
        messages.error(request, f'Error al descargar: {str(e)}')
        return redirect('reportes:detalle_reporte', pk=pk)

@login_required
def editar_reporte_view(request, pk):
    """Vista para editar los datos básicos de un reporte."""
    from .forms import EditarReporteForm
    reporte = get_object_or_404(Reporte, pk=pk)

    if request.method == 'POST':
        form = EditarReporteForm(request.POST, instance=reporte)
        if form.is_valid():
            form.save()
            messages.success(request, 'Reporte actualizado correctamente.')
            return redirect('reportes:detalle_reporte', pk=pk)
    else:
        form = EditarReporteForm(instance=reporte)

    return render(request, 'reportes/editar_reporte.html', {
        'form': form,
        'reporte': reporte,
    })


@login_required
@require_http_methods(["POST"])
def cancelar_reporte(request, pk):
    """Vista para cancelar/eliminar un reporte"""
    try:
        from usuarios.models import Usuario

        reporte = get_object_or_404(Reporte, id=pk)

        # Verificar permisos: solo el creador o staff puede cancelar
        try:
            usuario_actual = Usuario.objects.get(numero_documento=request.user.username)
            if reporte.usuario_solicitante != usuario_actual and not request.user.is_staff:
                return JsonResponse({
                    'success': False,
                    'error': 'No tienes permiso para cancelar este reporte'
                }, status=403)
        except Usuario.DoesNotExist:
            if not request.user.is_staff:
                return JsonResponse({
                    'success': False,
                    'error': 'No tienes permiso para cancelar este reporte'
                }, status=403)

        # Guardar información antes de eliminar
        titulo_reporte = reporte.titulo

        # Eliminar el reporte
        reporte.delete()

        return JsonResponse({
            'success': True,
            'message': f'Reporte "{titulo_reporte}" eliminado correctamente'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al cancelar reporte: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["POST"])
def eliminar_reporte(request, pk):
    """Elimina un reporte y redirige al dashboard."""
    from usuarios.models import Usuario
    reporte = get_object_or_404(Reporte, id=pk)
    try:
        usuario_actual = Usuario.objects.get(numero_documento=request.user.username)
        if reporte.usuario_solicitante != usuario_actual and not request.user.is_staff:
            messages.error(request, 'No tienes permiso para eliminar este reporte.')
            return redirect('reportes:dashboard')
    except Usuario.DoesNotExist:
        if not request.user.is_staff:
            messages.error(request, 'No tienes permiso para eliminar este reporte.')
            return redirect('reportes:dashboard')
    titulo = reporte.titulo
    reporte.delete()
    messages.success(request, f'Reporte "{titulo}" eliminado correctamente.')
    return redirect('reportes:dashboard')

@login_required
def tipos_reporte_view(request):
    return render(request, 'reportes/tipos_reporte.html', {'title': 'Tipos de Reporte'})

@login_required
def crear_tipo_reporte_view(request):
    return render(request, 'reportes/crear_tipo_reporte.html', {'title': 'Crear Tipo Reporte'})

@login_required
def metricas_view(request):
    return render(request, 'reportes/metricas.html', {'title': 'Métricas'})

@login_required
def kpis_view(request):
    return render(request, 'reportes/kpis.html', {'title': 'KPIs'})

@login_required
def reporte_eficiencia_view(request):
    return render(request, 'reportes/reporte_eficiencia.html', {'title': 'Reporte Eficiencia'})

@login_required
def reporte_costos_view(request):
    return render(request, 'reportes/reporte_costos.html', {'title': 'Reporte Costos'})

@login_required
def reporte_mantenimiento_view(request):
    return render(request, 'reportes/reporte_mantenimiento.html', {'title': 'Reporte Mantenimiento'})

@login_required
def analisis_tendencias_view(request):
    return render(request, 'reportes/analisis_tendencias.html', {'title': 'Análisis Tendencias'})

@login_required
def analisis_predicciones_view(request):
    return render(request, 'reportes/analisis_predicciones.html', {'title': 'Análisis Predicciones'})

@login_required
def analisis_comparativo_view(request):
    return render(request, 'reportes/analisis_comparativo.html', {'title': 'Análisis Comparativo'})

@login_required
def exportar_excel_view(request):
    return HttpResponse('Excel export placeholder', content_type='application/vnd.ms-excel')

@login_required
def exportar_pdf_view(request):
    return HttpResponse('PDF export placeholder', content_type='application/pdf')

@login_required
def exportar_csv_view(request):
    return HttpResponse('CSV export placeholder', content_type='text/csv')

@login_required
def datos_widget_api(request, widget_id):
    return JsonResponse({'data': [], 'widget_id': widget_id})

@login_required
def grafico_eficiencia_api(request):
    """API para gráfico de eficiencia de máquinas"""
    try:
        from maquinaria.models import Maquina, CategoriaMaquina

        # Eficiencia por categoría
        categorias_eficiencia = CategoriaMaquina.objects.annotate(
            avg_eficiencia=Avg('maquina__eficiencia')
        ).filter(avg_eficiencia__isnull=False)

        labels = []
        data = []
        colors = []

        for categoria in categorias_eficiencia:
            labels.append(categoria.nombre)
            data.append(round(categoria.avg_eficiencia, 1))
            colors.append(categoria.color)

        return JsonResponse({
            'labels': labels,
            'data': data,
            'colors': colors
        })

    except ImportError:
        return JsonResponse({'labels': [], 'data': [], 'colors': []})

@login_required
def grafico_costos_api(request):
    """API para gráfico de costos de mantenimiento"""
    try:
        from maquinaria.models import HistorialMaquina
        from django.db.models import Sum

        # Costos por mes (últimos 6 meses)
        hoy = date.today()
        labels = []
        data = []

        for i in range(6):
            mes_fecha = hoy - timedelta(days=30*i)
            mes_nombre = mes_fecha.strftime('%B')

            # Calcular costos del mes
            costos_mes = HistorialMaquina.objects.filter(
                fecha_evento__year=mes_fecha.year,
                fecha_evento__month=mes_fecha.month,
                costo_asociado__isnull=False
            ).aggregate(total=Sum('costo_asociado'))['total'] or 0

            labels.insert(0, mes_nombre)
            data.insert(0, float(costos_mes))

        return JsonResponse({
            'labels': labels,
            'data': data
        })

    except ImportError:
        return JsonResponse({'labels': [], 'data': []})

@login_required
def grafico_estados_api(request):
    """API para gráfico de estados de máquinas"""
    try:
        from maquinaria.models import Maquina

        estados_data = Maquina.objects.values('estado').annotate(
            count=Count('id')
        ).order_by('-count')

        labels = []
        data = []
        colors = ['#28a745', '#ffc107', '#dc3545', '#17a2b8', '#6c757d']

        for i, estado in enumerate(estados_data):
            labels.append(estado['estado'].replace('_', ' ').title())
            data.append(estado['count'])

        return JsonResponse({
            'labels': labels,
            'data': data,
            'colors': colors[:len(labels)]
        })

    except ImportError:
        return JsonResponse({'labels': [], 'data': [], 'colors': []})

@login_required
def tabla_maquinas_api(request):
    """API para tabla de máquinas con datos reales"""
    try:
        from maquinaria.models import Maquina

        maquinas = Maquina.objects.select_related('categoria', 'proveedor').all()[:20]

        data = []
        for maquina in maquinas:
            data.append({
                'codigo': maquina.codigo_inventario,
                'nombre': maquina.nombre,
                'categoria': maquina.categoria.nombre if maquina.categoria else '-',
                'estado': maquina.get_estado_display(),
                'condicion': maquina.get_condicion_display(),
                'eficiencia': f"{maquina.eficiencia}%",
                'horas_uso': maquina.horas_uso_total,
                'ultimo_mantenimiento': maquina.fecha_ultimo_mantenimiento.strftime('%Y-%m-%d') if maquina.fecha_ultimo_mantenimiento else '-',
            })

        return JsonResponse({'data': data})

    except ImportError:
        return JsonResponse({'data': []})

@login_required
def personalizar_dashboard_view(request):
    return render(request, 'reportes/personalizar_dashboard.html', {'title': 'Personalizar Dashboard'})

@login_required
@require_http_methods(["POST"])
def agregar_widget_dashboard(request):
    return JsonResponse({'success': True})

@login_required
@require_http_methods(["POST"])
def eliminar_widget_dashboard(request, pk):
    return JsonResponse({'success': True})

@login_required
def preview_reporte_api(request):
    """API que devuelve una vista previa de los datos que incluirá el reporte"""
    try:
        tipo_id = request.GET.get('tipo_reporte') or request.POST.get('tipo_reporte')
        categoria_ids = request.GET.getlist('categoria') or request.POST.getlist('categoria')
        categoria_id = categoria_ids[0] if len(categoria_ids) == 1 else None  # para filtros de nombre
        centro = request.GET.get('centro_formacion') or request.POST.get('centro_formacion')
        estado_maq = request.GET.get('estado_maquina') or request.POST.get('estado_maquina')
        fecha_inicio = request.GET.get('fecha_inicio') or request.POST.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin') or request.POST.get('fecha_fin')
        formato = request.GET.get('formato') or request.POST.get('formato', 'pdf')

        tipo_nombre = 'Reporte'
        tipo_lower = ''
        if tipo_id:
            try:
                tipo_obj = TipoReporte.objects.get(id=tipo_id)
                tipo_nombre = tipo_obj.nombre
                tipo_lower = tipo_nombre.lower()
            except TipoReporte.DoesNotExist:
                pass

        filtros_aplicados = []
        if centro:
            filtros_aplicados.append(f'Centro: {centro}')
        if fecha_inicio:
            filtros_aplicados.append(f'Desde: {fecha_inicio}')
        if fecha_fin:
            filtros_aplicados.append(f'Hasta: {fecha_fin}')

        # ---- MAQUINARIA ----
        if 'maquinaria' in tipo_lower:
            from maquinaria.models import Maquina, CategoriaMaquina, MantenimientoProgramado, UsoMaquinaria
            qs = Maquina.objects.select_related('categoria').all()
            if categoria_ids:
                qs = qs.filter(categoria_id__in=categoria_ids)
                nombres_cats = list(CategoriaMaquina.objects.filter(id__in=categoria_ids).values_list('nombre', flat=True))
                if nombres_cats:
                    filtros_aplicados.append(f'Categorías: {", ".join(nombres_cats)}')
            if centro:
                qs = qs.filter(centro_formacion=centro)
            if estado_maq:
                qs = qs.filter(estado=estado_maq)
                filtros_aplicados.append(f'Estado: {estado_maq}')
            if fecha_inicio:
                qs = qs.filter(fecha_adquisicion__gte=fecha_inicio)
            if fecha_fin:
                qs = qs.filter(fecha_adquisicion__lte=fecha_fin)

            total = qs.count()
            ids = list(qs.values_list('id', flat=True))
            total_mantenimientos = MantenimientoProgramado.objects.filter(maquina_id__in=ids).count()
            total_usos = UsoMaquinaria.objects.filter(maquina_id__in=ids).count()

            headers = ['Código', 'Nombre', 'Categoría', 'Estado', 'Condición', 'Eficiencia (%)', 'Centro']
            muestra = [[m.codigo_inventario, m.nombre,
                        m.categoria.nombre if m.categoria else '-',
                        m.get_estado_display(), m.get_condicion_display(),
                        str(m.eficiencia), m.centro_formacion]
                       for m in qs[:5]]

            estado_resumen = dict(qs.values_list('estado').annotate(c=Count('id')))

            return JsonResponse({
                'success': True,
                'tipo_nombre': tipo_nombre,
                'formato': formato.upper(),
                'total_registros': total,
                'resumen_secciones': [
                    {'nombre': 'Máquinas', 'total': total},
                    {'nombre': 'Mantenimientos', 'total': total_mantenimientos},
                    {'nombre': 'Usos registrados', 'total': total_usos},
                ],
                'columnas': headers,
                'muestra': muestra,
                'estado_resumen': estado_resumen,
                'filtros_aplicados': filtros_aplicados,
            })

        # ---- INVENTARIO ----
        elif 'inventario' in tipo_lower:
            from inventario.models import PiezaInventario
            qs = PiezaInventario.objects.all()
            if centro:
                qs = qs.filter(centro_formacion=centro)
            if fecha_inicio:
                qs = qs.filter(fecha_registro__date__gte=fecha_inicio)
            if fecha_fin:
                qs = qs.filter(fecha_registro__date__lte=fecha_fin)

            total = qs.count()
            condicion_resumen = dict(qs.values_list('condicion').annotate(c=Count('id')))
            headers = ['Código', 'Nombre', 'Categoría', 'Condición', 'Horas Uso', 'Fecha Registro', 'Centro']
            muestra = [[p.codigo_inventario, p.nombre,
                        p.categoria or '-', p.get_condicion_display(),
                        str(p.horas_uso),
                        p.fecha_registro.strftime('%Y-%m-%d') if p.fecha_registro else 'N/A',
                        p.centro_formacion or '-']
                       for p in qs.order_by('-fecha_registro')[:5]]

            return JsonResponse({
                'success': True,
                'tipo_nombre': tipo_nombre,
                'formato': formato.upper(),
                'total_registros': total,
                'resumen_secciones': [{'nombre': 'Piezas de Inventario', 'total': total}],
                'columnas': headers,
                'muestra': muestra,
                'estado_resumen': condicion_resumen,
                'filtros_aplicados': filtros_aplicados,
            })

        # ---- DOCUMENTOS ----
        elif 'documento' in tipo_lower:
            from documentos.models import Documento
            qs = Documento.objects.all()
            if fecha_inicio:
                qs = qs.filter(fecha_creacion__date__gte=fecha_inicio)
            if fecha_fin:
                qs = qs.filter(fecha_creacion__date__lte=fecha_fin)

            total = qs.count()
            por_tipo = list(qs.values('tipo_documento__nombre').annotate(c=Count('id'))
                            .order_by('-c').values_list('tipo_documento__nombre', 'c'))
            por_estado = dict(qs.values_list('estado').annotate(c=Count('id')))
            headers = ['Título', 'Tipo', 'Categoría', 'Estado', 'Fecha Creación']
            muestra = [[d.titulo, d.tipo_documento.nombre if d.tipo_documento else '-',
                        d.categoria.nombre if d.categoria else '-',
                        dict(Documento.ESTADO_CHOICES).get(d.estado, d.estado),
                        d.fecha_creacion.strftime('%Y-%m-%d') if d.fecha_creacion else 'N/A']
                       for d in qs.order_by('-fecha_creacion')[:5]]

            return JsonResponse({
                'success': True,
                'tipo_nombre': tipo_nombre,
                'formato': formato.upper(),
                'total_registros': total,
                'resumen_secciones': [
                    {'nombre': t[0] or 'Sin tipo', 'total': t[1]} for t in por_tipo[:5]
                ],
                'columnas': headers,
                'muestra': muestra,
                'estado_resumen': por_estado,
                'filtros_aplicados': filtros_aplicados,
            })

        else:
            return JsonResponse({
                'success': False,
                'error': 'Selecciona un tipo de reporte válido.'
            }, status=400)

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def filtros_centros_api(request):
    return JsonResponse({'centros': []})

@login_required
def filtros_categorias_api(request):
    return JsonResponse({'categorias': []})

@login_required
def aplicar_filtros_api(request):
    return JsonResponse({'success': True, 'data': []})

@login_required
def reportes_programados_view(request):
    return render(request, 'reportes/reportes_programados.html', {'title': 'Reportes Programados'})

@login_required
def crear_reporte_programado_view(request):
    return render(request, 'reportes/crear_reporte_programado.html', {'title': 'Crear Reporte Programado'})
